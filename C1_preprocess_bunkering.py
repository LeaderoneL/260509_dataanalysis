#!/usr/bin/env python3
"""
步骤 1：加油数据预处理 — 时间标准化与字段解析。

整合 PDF 需求 + 用户补充需求：
  - 处理双时间格式（datetime 和 "DD Mon YYYY HH:MM"）
  - 清洗 location_details 中的不间断空格（\\xa0）
  - 解析船舶信息、时长、吃水深度
  - 根据 vessel_raw 边界分配交易 ID
  - 对照锚地参考表进行港口匹配
  - 从 STS location_details 中提取供应商

输入：Bunkering Record v3.xlsx
输出：2_bunkering_preprocessed.xlsx
"""

import re
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ── 文件路径 ────────────────────────────────────────────────────────
# 所有路径均相对于此脚本所在位置（260509_dataanalysis/）
BASE_DIR = Path(__file__).resolve().parent
SRC = BASE_DIR / "Bunkering Record v3.xlsx"
PREPROCESSED = BASE_DIR / "2_bunkering_preprocessed.xlsx"

# ── 锚地参考表 ─────────────────────────────────────────────────────────
# 将英文位置关键词映射到中文港口简称。
# 顺序有影响：越靠前的条目匹配优先级越高。
# "Qushan" 重复出现（带和不带 "Bunkering" 后缀）以捕获
# 完整名称和被 Excel 截断的字符串。
ANCHORAGE_MAP = [
    ("Tide and Berth",      "秀山东"),   # 舟山 Tide and Berth 待泊锚地
    ("Xiazhimen South",     "条帚门"),   # 舟山虾峙门南锚地
    ("Xiazhimen North",     "虾峙门"),   # 舟山虾峙门北锚地
    ("Qushan Bunkering",    "衢山"),     # 舟山衢山加油锚地
    ("Qushan",              "衢山"),     # 截断变体（不含 "Bunkering"）
    ("Mashi",               "马峙"),     # 舟山马峙锚地
    ("Mazhi",               "马峙"),     # 替代罗马拼写
]

# ── 月份缩写 → 数字映射 ──────────────────────────────────────────
# 用于解析文本格式日期，如 "02 Feb 2023 08:11"
MONTH_ABBR = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}


def parse_datetime(val):
    """将两种原始时间格式统一解析为 datetime 对象。

    原始数据包含两种混合格式：
      格式 A — YYYY-MM-DD HH:MM:SS（例如 "2023-01-30 04:15:00"）
        以原生 datetime 或字符串形式存储，可选含秒数。
      格式 B — DD Mon YYYY HH:MM（例如 "02 Feb 2023 08:11"）
        以带缩写月份名的纯文本字符串形式存储。

    无法解析的值返回 None（调用处会触发警告）。
    """
    # 已经是 datetime 对象（pandas 会将部分单元格读取为原生类型）
    if isinstance(val, datetime):
        return val

    val = str(val).strip()
    if not val or val in ("None", "nan"):
        return None

    # 格式 A：类似 ISO 的时间戳，可选秒数
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2})(?::(\d{2}))?$", val)
    if m:
        y, mo, d, h, mi, s = m.groups()
        return datetime(int(y), int(mo), int(d), int(h), int(mi), int(s or 0))

    # 格式 B：带缩写月份的可读日期格式
    m = re.match(r"^(\d{1,2})\s+([A-Z][a-z]{2})\s+(\d{4})\s+(\d{2}):(\d{2})$", val)
    if m:
        d, mon_abbr, y, h, mi = m.groups()
        mo = MONTH_ABBR.get(mon_abbr)
        if mo:
            return datetime(int(y), mo, int(d), int(h), int(mi))

    print(f"  ⚠  无法解析 datetime：{val}")
    return None


def round_datetime(dt):
    """对小时进行四舍五入并调整 datetime，包括日期进位。

    依据第 1.4 节 v3 修正：当分钟 >= 30 时，小时 +1。
    若小时变为 24，则进位到次日 00:00 并更新日期。
    分钟/秒/微秒置零。
    dt 为 None 时返回 None。
    """
    if dt is None:
        return None
    if dt.minute >= 30:
        adj = dt + timedelta(hours=1)
    else:
        adj = dt
    return adj.replace(minute=0, second=0, microsecond=0)


def round_hour(adj_dt):
    """从已调整的 datetime 中提取小时数。"""
    if adj_dt is None:
        return None
    return adj_dt.hour


def fmt_date(dt):
    """将 datetime 格式化为 'DD Mon YYYY'（例如 '30 Jan 2023'）。

    符合 PDF 输出规范，与 Stata 日期格式惯例一致。
    """
    if dt is None:
        return None
    mmm = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
           7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}
    return f"{dt.day} {mmm[dt.month]} {dt.year}"


def parse_duration_hours(val):
    """将自由文本时长解析为总小时数（四舍五入）。

    支持混合单位：'X day(s)'、'X h'、'X min'。
    '-'（开始 == 结束）、空值或 null 值返回 0。
    分钟数四舍五入到最近的小时。
    """
    val = str(val).strip()
    if val in ("-", "", "None", "nan"):
        return 0.0
    total = 0.0
    # 天数 → 小时
    day_m = re.search(r"(\d+)\s*day[s]?", val)
    if day_m:
        total += int(day_m.group(1)) * 24
    # 整小时
    hour_m = re.search(r"(\d+)\s*h", val)
    if hour_m:
        total += int(hour_m.group(1))
    # 分钟 → 小时（四舍五入）
    min_m = re.search(r"(\d+)\s*min", val)
    if min_m:
        total += round(int(min_m.group(1)) / 60)
    return total


def parse_draught(val):
    """从类似 '-0.5 m' 或 '+0.3 m' 的字符串中提取吃水深度数值。

    返回找到的第一个带符号浮点数；无数字时返回 0.0。
    """
    val = str(val).strip()
    m = re.search(r"([+-]?\d+(?:\.\d+)?)", val)
    if m:
        return float(m.group(1))
    return 0.0


def extract_vessel_info(val):
    """将船舶信息字符串解析为结构化字段。

    预期格式：
      "VESSELNAME (CODE) - Type, Status, DWT dwt, GTgt"
    示例：
      "JOLLY ARGENTO (9467043) - Container Ship (Fully Cellular), In Service/Commission, 62510 dwt, 47789gt"

    返回包含 vessel_name、vessel_code、vessel_type、
    vessel_status、dwt、gt 键的字典。解析失败返回空字典。
    """
    val = str(val).strip()
    m = re.match(
        r"^(.+?)\s+\((\d+)\)\s*-\s*(.+?),\s*(.+?),\s*([\d,]+)\s*dwt,\s*([\d,]+)\s*gt$",
        val,
    )
    if m:
        return {
            "vessel_name":   m.group(1).strip(),
            "vessel_code":   m.group(2).strip(),
            "vessel_type":   m.group(3).strip(),
            "vessel_status": m.group(4).strip(),
            "dwt":           int(m.group(5).replace(",", "")),
            "gt":            int(m.group(6).replace(",", "")),
        }
    return {}


def extract_supplier(location_detail):
    """从 STS Location Details 中提取供应商（加油船）名称。

    STS 行格式为 "With <vessel name>"（例如 "With Fu Jie 168"）。
    首先清洗不间断空格（\\xa0）——当源报告使用非标准空白字符
    生成时会出现此类字符。
    """
    val = str(location_detail).strip()
    val = val.replace("\xa0", " ")       # 清洗不间断空格
    if val.lower().startswith("with "):
        return val[5:].strip()           # 去掉 "With " 前缀
    return val.strip()


def match_port(location_detail):
    """将锚地 Location Details 字符串与参考港口表进行匹配。

    原始数据常包含因 Excel 单元格宽度限制而被截断的字符串
    （例如 "At Zhoushan Tide and Berth W" 而非完整的
    "At Zhoushan Tide and Berth Waiting Anchorage"）。

    两级匹配：
      1) 完整子串匹配 —— 关键词完整出现在字符串中。
      2) 单词级截断匹配 —— 将关键词的每个单词与每个单词位置
         逐一比对，允许部分重叠。此方法可捕获 Excel 截断末尾单词
         的情况。要求所有关键词单词均匹配，或当到达位置字符串末尾时
         至少匹配 2 个单词。

    匹配成功返回中文港口名，无匹配返回 ""。
    """
    val = str(location_detail).strip()
    val = val.replace("\xa0", " ")       # 清洗不间断空格
    if val.lower().startswith("at "):
        val = val[3:]                    # 去掉 "At " 前缀
    val_lower = val.lower()
    val_words = val_lower.split()

    for keyword, port_name in ANCHORAGE_MAP:
        kw_lower = keyword.lower()

        # 第 1 级：完整子串匹配（最常见情况）
        if kw_lower in val_lower:
            return port_name

        # 第 2 级：针对截断字符串的单词级模糊匹配
        kw_words = kw_lower.split()
        for i in range(len(val_words)):
            matched = 0
            for j, kw_w in enumerate(kw_words):
                if i + j >= len(val_words):
                    break
                loc_w = val_words[i + j]
                # 任一单词是另一个的前缀（处理截断）
                if loc_w.startswith(kw_w) or kw_w.startswith(loc_w):
                    matched += 1
                else:
                    break
            # 所有关键词单词均匹配，或在字符串末尾部分匹配
            if matched == len(kw_words) or (matched >= 2 and i + matched == len(val_words)):
                return port_name

    return ""


def is_sts(op_str):
    """检测 STS（加油）操作。

    处理 'TS - Bunkering' 截断变体，即前导 'S' 被 Excel 列宽裁剪的情况。
    """
    if not op_str:
        return False
    ops = str(op_str).strip().upper()
    return "STS" in ops or "TS" in ops


def is_anchorage(op_str):
    """检测 Anchorage（等待）操作。"""
    if not op_str:
        return False
    return "ANCHORAGE" in str(op_str).strip().upper()


def main():
    """
    ┌─────────────────────────────────────────────────────────────┐
    │  处理流程概览                                                │
    │                                                             │
    │  原始 Excel（9 列，4523 行）                                  │
    │    → 根据 vessel_raw 边界分配 transaction_id                 │
    │    → 将船舶信息向前填充到空行                                  │
    │    → 解析船舶名称/代码/类型/状态/dwt/gt                        │
    │    → 统一双时间格式 → 标准 datetime                           │
    │    → 小时四舍五入并处理日期进位（第 1.4 节 v3 修正）            │
    │    → 解析时长字符串 → 数字小时                                 │
    │    → 解析吃水深度字符串 → 数值                                 │
    │    → 分类每行为 STS / Anchorage / 其他                        │
    │    → 从 STS 行提取供应商名称                                  │
    │    → 将锚地行与参考港口表匹配                                  │
    │    → 保存格式化的预处理 xlsx                                   │
    └─────────────────────────────────────────────────────────────┘
    """

    # ── 1. 读取原始数据 ───────────────────────────────────────────────
    # 第 0 行 = 报告元数据，第 1 行 = 列标题 → skiprows=2
    print("=" * 60)
    print("Step 1：读取原始数据")
    df = pd.read_excel(SRC, header=None, skiprows=2)
    df.columns = [
        "vessel_raw", "col_extra", "operation", "location", "country",
        "start_raw", "end_raw", "duration_raw", "draught_raw", "location_details",
    ]
    df = df.drop(columns=["col_extra"])  # v3 文件中 col 1 为空列，无实际数据
    print(f"  原始行数：{len(df)}")

    # ── 2. 分配交易 ID ──────────────────────────────────────────────
    # 一笔交易跨多行：第一行在第 0 列有船舶名称，
    # 后续行均为空，直到下一笔交易。
    # 每次 vessel_raw 非空时 ID 递增。
    print("\nStep 2：分配交易 ID")
    txn_id = 0
    txn_ids = []
    for _, row in df.iterrows():
        if pd.notna(row["vessel_raw"]) and str(row["vessel_raw"]).strip() not in ("", "None", "nan"):
            txn_id += 1
        txn_ids.append(txn_id)
    df["transaction_id"] = txn_ids
    print(f"  总交易数：{txn_id}")

    # ── 3. 向前填充船舶信息 ────────────────────────────────────
    # 船舶信息仅出现在每笔交易的第一行。
    # 向前填充将其复制到该交易内的所有后续行。
    print("\nStep 3：向前填充船舶信息")
    mask = df["vessel_raw"].notna() & (df["vessel_raw"].astype(str).str.strip() != "")
    df.loc[~mask, "vessel_raw"] = None        # 将空字符串视为 NaN
    df["vessel_raw"] = df["vessel_raw"].ffill()
    print(f"  填充后非空船舶行数：{df['vessel_raw'].notna().sum()}")

    # ── 4. 解析船舶信息 ───────────────────────────────────────────
    # 从自由文本船舶描述中提取结构化字段。
    print("\nStep 4：解析船舶信息")
    vessel_info_list = [extract_vessel_info(v) for v in df["vessel_raw"]]
    df_vessel = pd.DataFrame(vessel_info_list)
    df = pd.concat([df, df_vessel], axis=1)
    unique_vessels = df["vessel_name"].nunique()
    print(f"  唯一船舶数：{unique_vessels}")

    # ── 5. 标准化时间格式 ────────────────────────────────────
    # 这是核心预处理步骤：无论原始格式如何，
    # 将两种 datetime 格式统一为一致的 datetime 对象。
    print("\nStep 5：标准化时间格式")
    df["start_dt"] = df["start_raw"].apply(parse_datetime)
    df["end_dt"] = df["end_raw"].apply(parse_datetime)
    n_fail_start = df["start_dt"].isna().sum()
    n_fail_end = df["end_dt"].isna().sum()
    print(f"  解析失败 — start：{n_fail_start}，end：{n_fail_end}")

    # ── 6. 解析时长和吃水深度 ──────────────────────────────────
    print("\nStep 6：解析时长和吃水深度")
    df["duration_hours"] = df["duration_raw"].apply(parse_duration_hours)
    df["draught_numeric"] = df["draught_raw"].apply(parse_draught)
    print("  完成")

    # ── 7. 预计算补充字段 ────────────────────────────
    # 这些派生列直接供最终化步骤使用，
    # 无需重新解析原始字符串。
    print("\nStep 7：预计算补充字段")
    df["is_sts"] = df["operation"].apply(is_sts)
    df["is_anchorage"] = df["operation"].apply(is_anchorage)
    df["supplier"] = df.apply(
        lambda r: extract_supplier(r["location_details"]) if r["is_sts"] else "",
        axis=1,
    )
    df["port_matched"] = df.apply(
        lambda r: match_port(r["location_details"]) if r["is_anchorage"] else "",
        axis=1,
    )
    # 带小时四舍五入和日期进位的调整后 datetime（第 1.4 节 v3 修正）
    df["start_adj"] = df["start_dt"].apply(round_datetime)
    df["end_adj"] = df["end_dt"].apply(round_datetime)
    # 从调整后 datetime 派生的预格式化日期/小时列
    df["start_fmt"] = df["start_adj"].apply(fmt_date)
    df["end_fmt"] = df["end_adj"].apply(fmt_date)
    df["start_hour"] = df["start_adj"].apply(round_hour)
    df["end_hour"] = df["end_adj"].apply(round_hour)

    n_sts = df["is_sts"].sum()
    n_anch = df["is_anchorage"].sum()
    print(f"  STS 行数：{n_sts}，Anchorage 行数：{n_anch}")

    # ── 8. 保存预处理后的 Excel ─────────────────────────────────────
    # 使用 openpyxl 进行单元格级格式化（粗体标题、蓝色背景、
    # 列宽）。所有原始 + 解析字段均被保存，以便最终化
    # 步骤直接读取干净数据，无需重新运行提取器。
    print(f"\nStep 8：保存预处理文件 → {PREPROCESSED}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "预处理数据"

    headers = [
        "row_num", "transaction_id",
        "vessel_name", "vessel_code", "vessel_type", "vessel_status", "dwt", "gt",
        "operation", "location", "country",
        "start_original", "end_original",
        "start_datetime", "end_datetime",
        "start_adjusted", "end_adjusted",
        "start_formatted", "end_formatted",
        "start_hour", "end_hour",
        "duration_original", "duration_hours",
        "draught_original", "draught_numeric",
        "is_sts", "is_anchorage",
        "supplier", "port_matched",
        "location_details",
    ]

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, (_, row) in enumerate(df.iterrows()):
        r = i + 2
        vals = [
            i + 3,          # 原始 Excel 行号（从 1 开始计数；+2 是因为跳过了两行）
            row["transaction_id"],
            row.get("vessel_name"),
            row.get("vessel_code"),
            row.get("vessel_type"),
            row.get("vessel_status"),
            row.get("dwt"),
            row.get("gt"),
            row["operation"],
            row["location"],
            row["country"],
            str(row["start_raw"]) if pd.notna(row["start_raw"]) else None,
            str(row["end_raw"]) if pd.notna(row["end_raw"]) else None,
            row["start_dt"].strftime("%Y-%m-%d %H:%M:%S") if pd.notna(row["start_dt"]) and row["start_dt"] else None,
            row["end_dt"].strftime("%Y-%m-%d %H:%M:%S") if pd.notna(row["end_dt"]) and row["end_dt"] else None,
            row["start_adj"].strftime("%Y-%m-%d %H:%M:%S") if pd.notna(row["start_adj"]) and row["start_adj"] else None,
            row["end_adj"].strftime("%Y-%m-%d %H:%M:%S") if pd.notna(row["end_adj"]) and row["end_adj"] else None,
            row["start_fmt"],
            row["end_fmt"],
            row["start_hour"],
            row["end_hour"],
            str(row["duration_raw"]) if pd.notna(row["duration_raw"]) else None,
            row["duration_hours"],
            str(row["draught_raw"]) if pd.notna(row["draught_raw"]) else None,
            row["draught_numeric"],
            1 if row["is_sts"] else 0,
            1 if row["is_anchorage"] else 0,
            row["supplier"] if row["supplier"] else None,
            row["port_matched"] if row["port_matched"] else None,
            str(row["location_details"]).replace("\xa0", " ") if pd.notna(row["location_details"]) else None,
        ]
        for c, val in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=val)

    # 为可读性调整的列宽
    col_widths = {1: 8, 2: 14, 3: 30, 4: 12, 5: 30, 6: 24, 7: 10, 8: 10,
                  9: 16, 10: 14, 11: 8, 12: 22, 13: 22, 14: 20, 15: 20,
                  16: 20, 17: 20, 18: 16, 19: 16, 20: 8, 21: 8,
                  22: 18, 23: 14, 24: 14, 25: 14, 26: 8, 27: 14,
                  28: 12, 29: 14, 30: 50}
    for col_idx, width in col_widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    wb.save(PREPROCESSED)
    print(f"  已保存 {len(df)} 行 × {len(headers)} 列")

    # ── 9. 汇总统计 ──────────────────────────────────────────
    print("\n" + "=" * 60)
    print("预处理汇总：")
    print(f"  总行数：              {len(df)}")
    print(f"  总交易数：            {df['transaction_id'].nunique()}")
    print(f"  唯一船舶数：          {df['vessel_name'].nunique()}")
    print(f"  STS 行数：            {df['is_sts'].sum()}")
    print(f"  Anchorage 行数：      {df['is_anchorage'].sum()}")
    print(f"  其他操作行数：        {len(df) - df['is_sts'].sum() - df['is_anchorage'].sum()}")
    # 港口匹配覆盖率 = 已匹配锚地行数 / 总锚地行数
    n_matched = df[df["port_matched"] != ""]["port_matched"].notna().sum()
    print(f"  港口匹配覆盖率：      {n_matched}/{df['is_anchorage'].sum()} 锚地行")
    print(f"  输出文件：            {PREPROCESSED}")
    print("完成！")


if __name__ == "__main__":
    main()
