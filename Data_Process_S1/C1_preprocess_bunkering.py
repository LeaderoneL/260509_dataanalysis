#!/usr/bin/env python3
"""
Step 1: Bunkering data preprocessing — time standardization and field parsing.

Integrates PDF requirements + user supplementary requirements:
  - Handles dual time formats (datetime and "DD Mon YYYY HH:MM")
  - Cleans non-breaking spaces (\\xa0) from location_details
  - Parses vessel info, duration, draught
  - Assigns transaction IDs from vessel_raw boundaries
  - Port matching against anchorage reference table
  - Supplier extraction from STS location_details

Input:  1_Bunkering_Record_revised_transaction.xlsx
Output: 2_bunkering_preprocessed.xlsx
"""

import re
import pandas as pd
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ── File paths ────────────────────────────────────────────────────────
# All paths are relative to this script's location (260509_dataanalysis/)
BASE_DIR = Path(__file__).resolve().parent
SRC = BASE_DIR / "1_Bunkering_Record_revised_transaction.xlsx"
PREPROCESSED = BASE_DIR / "2_bunkering_preprocessed.xlsx"

# ── Anchorage reference table ─────────────────────────────────────────
# Maps English location keywords to Chinese short port names.
# Order matters: earlier entries have higher matching priority.
# "Qushan" is duplicated (with and without "Bunkering" suffix) to catch
# both full names and truncated Excel strings.
ANCHORAGE_MAP = [
    ("Tide and Berth",      "秀山东"),   # Zhoushan Tide and Berth Waiting Anchorage
    ("Xiazhimen South",     "条帚门"),   # Zhoushan Xiazhimen South Anchorage
    ("Xiazhimen North",     "虾峙门"),   # Zhoushan Xiazhimen North Anchorage
    ("Qushan Bunkering",    "衢山"),     # Zhoushan Qushan Bunkering Anchorage
    ("Qushan",              "衢山"),     # Truncated variant without "Bunkering"
    ("Mashi",               "马峙"),     # Zhoushan Mashi Anchorage
    ("Mazhi",               "马峙"),     # Alternate romanization
]

# ── Month name → number mapping ──────────────────────────────────────
# Used to parse text-format dates like "02 Feb 2023 08:11"
MONTH_ABBR = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}


def parse_datetime(val):
    """Parse both source time formats into a unified datetime object.

    The raw data contains two intermixed formats:
      Format A — YYYY-MM-DD HH:MM:SS  (e.g. "2023-01-30 04:15:00")
        Stored as native datetime or string, optionally with seconds.
      Format B — DD Mon YYYY HH:MM      (e.g. "02 Feb 2023 08:11")
        Stored as a plain string with abbreviated month name.

    Returns None for unparseable values (triggers a warning at the call site).
    """
    # Already a datetime object (pandas reads some cells natively)
    if isinstance(val, datetime):
        return val

    val = str(val).strip()
    if not val or val in ("None", "nan"):
        return None

    # Format A: ISO-like timestamp with optional seconds
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2})(?::(\d{2}))?$", val)
    if m:
        y, mo, d, h, mi, s = m.groups()
        return datetime(int(y), int(mo), int(d), int(h), int(mi), int(s or 0))

    # Format B: human-readable date with abbreviated month
    m = re.match(r"^(\d{1,2})\s+([A-Z][a-z]{2})\s+(\d{4})\s+(\d{2}):(\d{2})$", val)
    if m:
        d, mon_abbr, y, h, mi = m.groups()
        mo = MONTH_ABBR.get(mon_abbr)
        if mo:
            return datetime(int(y), mo, int(d), int(h), int(mi))

    print(f"  ⚠  Failed to parse datetime: {val}")
    return None


def round_hour(dt):
    """Round a datetime's hour to the nearest integer (>=30 min rounds up).

    24:00 wraps to 0 per the PDF specification.
    Returns None when dt is None (unparseable source value).
    """
    if dt is None:
        return None
    if dt.minute >= 30:
        h = dt.hour + 1
        return 0 if h == 24 else h
    return dt.hour


def fmt_date(dt):
    """Format a datetime as 'DD Mon YYYY' (e.g. '30 Jan 2023').

    This matches the PDF output spec and is consistent with Stata
    date formatting conventions.
    """
    if dt is None:
        return None
    mmm = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
           7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}
    return f"{dt.day} {mmm[dt.month]} {dt.year}"


def parse_duration_hours(val):
    """Parse a free-text duration into total hours (rounded).

    Handles mixed units: 'X day(s)', 'X h', 'X min'.
    Returns 0 for '-' (start == end), empty, or null values.
    Minutes are rounded to the nearest hour.
    """
    val = str(val).strip()
    if val in ("-", "", "None", "nan"):
        return 0.0
    total = 0.0
    # Days → hours
    day_m = re.search(r"(\d+)\s*day[s]?", val)
    if day_m:
        total += int(day_m.group(1)) * 24
    # Whole hours
    hour_m = re.search(r"(\d+)\s*h", val)
    if hour_m:
        total += int(hour_m.group(1))
    # Minutes → hours (rounded)
    min_m = re.search(r"(\d+)\s*min", val)
    if min_m:
        total += round(int(min_m.group(1)) / 60)
    return total


def parse_draught(val):
    """Extract the numeric draught value from a string like '-0.5 m' or '+0.3 m'.

    Returns the first signed float found; 0.0 if no number is present.
    """
    val = str(val).strip()
    m = re.search(r"([+-]?\d+(?:\.\d+)?)", val)
    if m:
        return float(m.group(1))
    return 0.0


def extract_vessel_info(val):
    """Parse a vessel info string into structured fields.

    Expected format:
      "VESSELNAME (CODE) - Type, Status, DWT dwt, GTgt"
    Example:
      "JOLLY ARGENTO (9467043) - Container Ship (Fully Cellular), In Service/Commission, 62510 dwt, 47789gt"

    Returns a dict with keys: vessel_name, vessel_code, vessel_type,
    vessel_status, dwt, gt.  Returns an empty dict on parse failure.
    """
    val = str(val).strip()
    m = re.match(
        r"^(.+?)\s+\((\d+)\)\s*-\s*(.+?),\s*(.+?),\s*([\d,]+)\s*dwt,\s*([\d,]+)\s*gt$",
        val,
    )
    if m:
        return {
            "vessel_name":   m.group(1).strip(),
            "vessel_code":   m.group(2).strip(),
            "vessel_type":   m.group(3).strip(),
            "vessel_status": m.group(4).strip(),
            "dwt":           int(m.group(5).replace(",", "")),
            "gt":            int(m.group(6).replace(",", "")),
        }
    return {}


def extract_supplier(location_detail):
    """Extract the supplier (bunkering vessel) name from STS Location Details.

    STS rows have format "With <vessel name>" (e.g. "With Fu Jie 168").
    Non-breaking spaces (\\xa0) are cleaned first — they appear when
    the source report was generated with non-standard whitespace.
    """
    val = str(location_detail).strip()
    val = val.replace("\xa0", " ")       # Clean non-breaking space
    if val.lower().startswith("with "):
        return val[5:].strip()           # Strip "With " prefix
    return val.strip()


def match_port(location_detail):
    """Match an anchorage Location Details string to the reference port table.

    The raw data often contains Excel-cell-width-truncated strings
    (e.g. "At Zhoushan Tide and Berth W" instead of the full
    "At Zhoushan Tide and Berth Waiting Anchorage").

    Two-tier matching:
      1) Full substring match — the keyword appears intact within the string.
      2) Word-level truncated match — each keyword word is checked
         against each word position, allowing partial overlap.  This
         catches cases where Excel truncated the last word(s).  A match
         requires either all keyword words to match, or at least 2
         words to match when we reach the end of the location string.

    Returns the Chinese port name on match, or "" if no match.
    """
    val = str(location_detail).strip()
    val = val.replace("\xa0", " ")       # Clean non-breaking space
    if val.lower().startswith("at "):
        val = val[3:]                    # Strip "At " prefix
    val_lower = val.lower()
    val_words = val_lower.split()

    for keyword, port_name in ANCHORAGE_MAP:
        kw_lower = keyword.lower()

        # Tier 1: exact substring match (most common case)
        if kw_lower in val_lower:
            return port_name

        # Tier 2: word-level fuzzy match for truncated strings
        kw_words = kw_lower.split()
        for i in range(len(val_words)):
            matched = 0
            for j, kw_w in enumerate(kw_words):
                if i + j >= len(val_words):
                    break
                loc_w = val_words[i + j]
                # Either word is a prefix of the other (handles truncation)
                if loc_w.startswith(kw_w) or kw_w.startswith(loc_w):
                    matched += 1
                else:
                    break
            # All keyword words matched, or partial match hitting the end
            if matched == len(kw_words) or (matched >= 2 and i + matched == len(val_words)):
                return port_name

    return ""


def is_sts(op_str):
    """Detect STS (bunkering) operations.

    Handles the 'TS - Bunkering' truncation variant where the leading 'S'
    was clipped by Excel column width.
    """
    if not op_str:
        return False
    ops = str(op_str).strip().upper()
    return "STS" in ops or "TS" in ops


def is_anchorage(op_str):
    """Detect Anchorage (waiting) operations."""
    if not op_str:
        return False
    return "ANCHORAGE" in str(op_str).strip().upper()


def main():
    """
    ┌─────────────────────────────────────────────────────────────┐
    │  PIPELINE OVERVIEW                                          │
    │                                                             │
    │  Raw Excel (9 cols, 4523 rows)                              │
    │    → Assign transaction_id from vessel_raw boundaries       │
    │    → Forward-fill vessel info into empty rows                │
    │    → Parse vessel name/code/type/status/dwt/gt              │
    │    → Standardize dual time formats → unified datetime        │
    │    → Parse duration string → numeric hours                   │
    │    → Parse draught string → numeric value                    │
    │    → Classify each row as STS / Anchorage / Other           │
    │    → Extract supplier name from STS rows                    │
    │    → Match anchorage rows to reference port table           │
    │    → Save formatted preprocessed xlsx                       │
    └─────────────────────────────────────────────────────────────┘
    """

    # ── 1. Read raw data ───────────────────────────────────────────────
    # Row 0 = report metadata, Row 1 = column headers → skiprows=2
    print("=" * 60)
    print("Step 1: Reading raw data")
    df = pd.read_excel(SRC, header=None, skiprows=2)
    df.columns = [
        "vessel_raw", "operation", "location", "country",
        "start_raw", "end_raw", "duration_raw", "draught_raw", "location_details",
    ]
    print(f"  Raw rows: {len(df)}")

    # ── 2. Assign transaction IDs ──────────────────────────────────────
    # A transaction spans multiple rows: the first row has a vessel name
    # in col 0, and subsequent rows are empty until the next transaction.
    # We increment the ID each time vessel_raw is non-empty.
    print("\nStep 2: Assigning transaction IDs")
    txn_id = 0
    txn_ids = []
    for _, row in df.iterrows():
        if pd.notna(row["vessel_raw"]) and str(row["vessel_raw"]).strip() not in ("", "None", "nan"):
            txn_id += 1
        txn_ids.append(txn_id)
    df["transaction_id"] = txn_ids
    print(f"  Total transactions: {txn_id}")

    # ── 3. Forward-fill vessel info ────────────────────────────────────
    # Vessel info only appears on the first row of each transaction.
    # Forward-fill copies it to all subsequent rows within that transaction.
    print("\nStep 3: Forward filling vessel info")
    mask = df["vessel_raw"].notna() & (df["vessel_raw"].astype(str).str.strip() != "")
    df.loc[~mask, "vessel_raw"] = None        # Treat empty strings as NaN
    df["vessel_raw"] = df["vessel_raw"].ffill()
    print(f"  Non-null vessel rows after fill: {df['vessel_raw'].notna().sum()}")

    # ── 4. Parse vessel info ───────────────────────────────────────────
    # Extract structured fields from the free-text vessel description.
    print("\nStep 4: Parsing vessel info")
    vessel_info_list = [extract_vessel_info(v) for v in df["vessel_raw"]]
    df_vessel = pd.DataFrame(vessel_info_list)
    df = pd.concat([df, df_vessel], axis=1)
    unique_vessels = df["vessel_name"].nunique()
    print(f"  Unique vessels: {unique_vessels}")

    # ── 5. Standardize time formats ────────────────────────────────────
    # This is the core preprocessing step: unify both datetime formats
    # into consistent datetime objects regardless of source format.
    print("\nStep 5: Standardizing time formats")
    df["start_dt"] = df["start_raw"].apply(parse_datetime)
    df["end_dt"] = df["end_raw"].apply(parse_datetime)
    n_fail_start = df["start_dt"].isna().sum()
    n_fail_end = df["end_dt"].isna().sum()
    print(f"  Parse failures — start: {n_fail_start}, end: {n_fail_end}")

    # ── 6. Parse duration and draught ──────────────────────────────────
    print("\nStep 6: Parsing duration and draught")
    df["duration_hours"] = df["duration_raw"].apply(parse_duration_hours)
    df["draught_numeric"] = df["draught_raw"].apply(parse_draught)
    print("  Done")

    # ── 7. Pre-compute supplementary fields ────────────────────────────
    # These derived columns feed directly into the finalization step
    # without re-parsing raw strings.
    print("\nStep 7: Pre-computing supplementary fields")
    df["is_sts"] = df["operation"].apply(is_sts)
    df["is_anchorage"] = df["operation"].apply(is_anchorage)
    df["supplier"] = df.apply(
        lambda r: extract_supplier(r["location_details"]) if r["is_sts"] else "",
        axis=1,
    )
    df["port_matched"] = df.apply(
        lambda r: match_port(r["location_details"]) if r["is_anchorage"] else "",
        axis=1,
    )
    # Pre-formatted date/hour columns for the final output spec
    df["start_fmt"] = df["start_dt"].apply(fmt_date)
    df["end_fmt"] = df["end_dt"].apply(fmt_date)
    df["start_hour"] = df["start_dt"].apply(round_hour)
    df["end_hour"] = df["end_dt"].apply(round_hour)

    n_sts = df["is_sts"].sum()
    n_anch = df["is_anchorage"].sum()
    print(f"  STS rows: {n_sts}, Anchorage rows: {n_anch}")

    # ── 8. Save preprocessed Excel ─────────────────────────────────────
    # Uses openpyxl for cell-level formatting (bold headers, blue background,
    # column widths).  All raw + parsed fields are saved so the finalization
    # step reads clean data without re-running the extractors.
    print(f"\nStep 8: Saving preprocessed file → {PREPROCESSED}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "预处理数据"

    headers = [
        "row_num", "transaction_id",
        "vessel_name", "vessel_code", "vessel_type", "vessel_status", "dwt", "gt",
        "operation", "location", "country",
        "start_original", "end_original",
        "start_datetime", "end_datetime",
        "start_formatted", "end_formatted",
        "start_hour", "end_hour",
        "duration_original", "duration_hours",
        "draught_original", "draught_numeric",
        "is_sts", "is_anchorage",
        "supplier", "port_matched",
        "location_details",
    ]

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, (_, row) in enumerate(df.iterrows()):
        r = i + 2
        vals = [
            i + 3,          # Original Excel row number (1-indexed; +2 for the two skipped rows)
            row["transaction_id"],
            row.get("vessel_name"),
            row.get("vessel_code"),
            row.get("vessel_type"),
            row.get("vessel_status"),
            row.get("dwt"),
            row.get("gt"),
            row["operation"],
            row["location"],
            row["country"],
            str(row["start_raw"]) if pd.notna(row["start_raw"]) else None,
            str(row["end_raw"]) if pd.notna(row["end_raw"]) else None,
            row["start_dt"].strftime("%Y-%m-%d %H:%M:%S") if pd.notna(row["start_dt"]) and row["start_dt"] else None,
            row["end_dt"].strftime("%Y-%m-%d %H:%M:%S") if pd.notna(row["end_dt"]) and row["end_dt"] else None,
            row["start_fmt"],
            row["end_fmt"],
            row["start_hour"],
            row["end_hour"],
            str(row["duration_raw"]) if pd.notna(row["duration_raw"]) else None,
            row["duration_hours"],
            str(row["draught_raw"]) if pd.notna(row["draught_raw"]) else None,
            row["draught_numeric"],
            1 if row["is_sts"] else 0,
            1 if row["is_anchorage"] else 0,
            row["supplier"] if row["supplier"] else None,
            row["port_matched"] if row["port_matched"] else None,
            str(row["location_details"]).replace("\xa0", " ") if pd.notna(row["location_details"]) else None,
        ]
        for c, val in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths tuned for readability
    col_widths = {1: 8, 2: 14, 3: 30, 4: 12, 5: 30, 6: 24, 7: 10, 8: 10,
                  9: 16, 10: 14, 11: 8, 12: 22, 13: 22, 14: 20, 15: 20,
                  16: 16, 17: 16, 18: 8, 19: 8, 20: 18, 21: 14, 22: 14, 23: 14,
                  24: 8, 25: 14, 26: 12, 27: 14, 28: 50}
    for col_idx, width in col_widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    wb.save(PREPROCESSED)
    print(f"  Saved {len(df)} rows × {len(headers)} columns")

    # ── 9. Summary statistics ──────────────────────────────────────────
    print("\n" + "=" * 60)
    print("Preprocessing summary:")
    print(f"  Total rows:             {len(df)}")
    print(f"  Total transactions:     {df['transaction_id'].nunique()}")
    print(f"  Unique vessels:         {df['vessel_name'].nunique()}")
    print(f"  STS rows:               {df['is_sts'].sum()}")
    print(f"  Anchorage rows:         {df['is_anchorage'].sum()}")
    print(f"  Other operation rows:   {len(df) - df['is_sts'].sum() - df['is_anchorage'].sum()}")
    # Port match coverage = matched anchorage rows / total anchorage rows
    n_matched = df[df["port_matched"] != ""]["port_matched"].notna().sum()
    print(f"  Port match coverage:    {n_matched}/{df['is_anchorage'].sum()} anchorage rows")
    print(f"  Output:                 {PREPROCESSED}")
    print("Done!")


if __name__ == "__main__":
    main()
