#!/usr/bin/env python3
"""
Step 2: Bunkering data finalization — transaction-level aggregation.

Reads the preprocessed xlsx and produces one row per transaction:
  - Vessel info: name, code, type, status, dwt, gt
  - Overall timeline: startdate, starthour, enddate, endhour, duration (hours)
  - Per-STS expansion: supplier{N}, start_STS{N}, starthour_STS{N},
    end_STS{N}, endhour_STS{N}, duration_STS{N}  (N = 1..max_sts)
  - STS total: duration_STS  (sum of all per-STS durations)
  - Port: selected from anchorage rows (prefers matched reference port)
  - unmatched_port: raw anchorage location_details when no port matched
  - Draught: mean numeric draught across all rows in the transaction
  - Deduplication + cross-validation against preprocessed data

Input:  2_bunkering_preprocessed.xlsx
Output: 3_bunkering_final.xlsx
"""

import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ── File paths ────────────────────────────────────────────────────────
# All paths are relative to this script's location (260509_dataanalysis/)
BASE_DIR = Path(__file__).resolve().parent
PREPROCESSED = BASE_DIR / "2_bunkering_preprocessed.xlsx"
FINAL = BASE_DIR / "3_bunkering_final.xlsx"


def duration_hours_between(start_dt, end_dt, fallback=0):
    """Compute elapsed hours between two datetimes (rounded to nearest hour).

    Falls back to the pre-parsed duration when either datetime is missing.
    This handles edge cases where STS start==end (raw duration "-"),
    yielding 0 hours from the datetime diff.
    """
    if pd.notna(start_dt) and pd.notna(end_dt) and start_dt and end_dt:
        return round((end_dt - start_dt).total_seconds() / 3600)
    return int(fallback)


def main():
    """
    ┌─────────────────────────────────────────────────────────────┐
    │  PIPELINE OVERVIEW                                          │
    │                                                             │
    │  Preprocessed xlsx (4523 rows × 28 cols)                    │
    │    → Read + re-parse datetime columns for grouping          │
    │    → Compute max_sts across all transactions                │
    │    → For each transaction:                                  │
    │        • Extract overall start / end / duration             │
    │        • Expand each STS into numbered columns              │
    │        • Pick port from anchorage (prefer matched ref)      │
    │        • Collect unmatched location_details                 │
    │        • Compute mean draught                               │
    │    → Remove duplicate transaction rows                      │
    │    → Save formatted final xlsx with dynamic STS columns     │
    │    → Validate: sample comparison + global statistics        │
    └─────────────────────────────────────────────────────────────┘
    """

    # ── 1. Read preprocessed data ──────────────────────────────────────
    print("=" * 60)
    print("Step 1: Reading preprocessed data")
    df = pd.read_excel(PREPROCESSED, sheet_name="预处理数据")
    print(f"  Rows: {len(df)}, Columns: {len(df.columns)}")

    # ── 2. Re-parse datetime strings into pandas Timestamp ─────────────
    # The preprocessed file stores datetimes as "YYY-MM-DD HH:MM:SS"
    # strings.  We convert them back to Timestamp for min/max/diff ops.
    print("\nStep 2: Parsing datetime columns")
    df["start_dt"] = pd.to_datetime(df["start_datetime"], errors="coerce")
    df["end_dt"] = pd.to_datetime(df["end_datetime"], errors="coerce")
    n_fail = df["start_dt"].isna().sum()
    print(f"  Datetime parse failures: {n_fail}")

    # ── 3. Group and determine dynamic column count ────────────────────
    # The number of STS columns depends on the transaction with the most
    # UNIQUE suppliers (not raw STS row count).  Same supplier doing
    # multiple deliveries is aggregated into one slot.
    print("\nStep 3: Grouping by transaction_id")
    txns = df.groupby("transaction_id", sort=True)

    max_sts = 0
    for tid, grp in txns:
        sts = grp[grp["is_sts"] == 1]
        unique_supp = sts["supplier"].dropna()
        unique_supp = unique_supp[unique_supp.astype(str).str.strip() != ""]
        unique_count = unique_supp.nunique()
        if unique_count > max_sts:
            max_sts = unique_count
    print(f"  Max unique suppliers per transaction: {max_sts}")

    # ── 4. Build transaction-level rows ────────────────────────────────
    print("\nStep 4: Building transaction-level rows")
    rows = []

    for tid, grp in txns:
        # Sort chronologically so min/max and positional ordering are
        # consistent within the transaction
        grp = grp.sort_values("start_dt")

        # ── Vessel fields ─────────────────────────────────────────────
        # Vessel info is identical across all rows in a transaction
        # (already forward-filled in preprocessing).  Take from first row.
        first = grp.iloc[0]
        vessel_fields = {
            "vessel_name":   first.get("vessel_name", ""),
            "vessel_code":   first.get("vessel_code", ""),
            "vessel_type":   first.get("vessel_type", ""),
            "vessel_status": first.get("vessel_status", ""),
            "dwt":           first.get("dwt", ""),
            "gt":            first.get("gt", ""),
        }

        # ── Overall start / end ────────────────────────────────────────
        # Uses ALL rows (STS + Anchorage + Other) to capture the full
        # transaction window.  The earliest start and latest end define
        # the transaction boundaries.
        overall_start = grp["start_dt"].min()
        overall_end   = grp["end_dt"].max()
        duration_total = duration_hours_between(overall_start, overall_end)

        # Retrieve the pre-formatted date string and rounded hour
        # from the row with the min start and max end respectively.
        # Guard against all-NaT groups (should not happen with clean data).
        start_idx = grp["start_dt"].idxmin() if overall_start is not pd.NaT else grp.index[0]
        end_idx   = grp["end_dt"].idxmax()   if overall_end is not pd.NaT   else grp.index[0]
        startdate = grp.loc[start_idx, "start_formatted"]
        starthour = grp.loc[start_idx, "start_hour"]
        enddate   = grp.loc[end_idx, "end_formatted"]
        endhour   = grp.loc[end_idx, "end_hour"]

        # ── STS expansion (grouped by UNIQUE supplier) ────────────────
        # Rows with the same supplier (e.g. same bunkering vessel
        # making multiple deliveries) are aggregated into one slot:
        #   start = earliest STS start
        #   end   = latest STS end
        #   duration = sum of all per-delivery durations
        # Supplier slots are ordered by first chronological appearance.
        sts_rows = grp[grp["is_sts"] == 1].sort_values("start_dt")

        # Group by unique supplier name, preserving first-appearance order
        supplier_order = []
        supplier_groups = {}
        for _, srow in sts_rows.iterrows():
            supp_raw = srow.get("supplier")
            supp = str(supp_raw).strip() if pd.notna(supp_raw) and str(supp_raw).strip() not in ("", "nan") else ""
            if not supp:
                continue
            if supp not in supplier_groups:
                supplier_groups[supp] = []
                supplier_order.append(supp)
            supplier_groups[supp].append(srow)

        sts_fields = {}
        sts_total_duration = 0.0
        for i, supp_name in enumerate(supplier_order, start=1):
            deliveries = supplier_groups[supp_name]
            # Aggregate across all deliveries of this supplier
            starts = [d["start_dt"] for d in deliveries if pd.notna(d["start_dt"])]
            ends   = [d["end_dt"]   for d in deliveries if pd.notna(d["end_dt"])]
            agg_start = min(starts) if starts else None
            agg_end   = max(ends)   if ends   else None

            # Total duration = sum of per-delivery durations
            total_dur = 0.0
            for d in deliveries:
                dur = duration_hours_between(
                    d["start_dt"], d["end_dt"],
                    fallback=d.get("duration_hours", 0),
                )
                total_dur += dur

            # Use the earliest delivery row for formatted date/hour
            min_idx = deliveries[0]["start_dt"]  # already sorted by start_dt
            # Find the delivery with min start_dt for start info
            start_delivery = min(deliveries, key=lambda d: d["start_dt"] if pd.notna(d["start_dt"]) else pd.Timestamp.max)
            end_delivery   = max(deliveries, key=lambda d: d["end_dt"]   if pd.notna(d["end_dt"])   else pd.Timestamp.min)

            sts_fields[f"supplier{i}"]      = supp_name
            sts_fields[f"start_STS{i}"]     = str(start_delivery["start_formatted"]) if pd.notna(start_delivery.get("start_formatted")) else ""
            sts_fields[f"starthour_STS{i}"] = int(start_delivery["start_hour"]) if pd.notna(start_delivery.get("start_hour")) else ""
            sts_fields[f"end_STS{i}"]       = str(end_delivery["end_formatted"]) if pd.notna(end_delivery.get("end_formatted")) else ""
            sts_fields[f"endhour_STS{i}"]   = int(end_delivery["end_hour"]) if pd.notna(end_delivery.get("end_hour")) else ""
            sts_fields[f"duration_STS{i}"]  = round(total_dur)
            sts_total_duration += total_dur

        # Number of unique suppliers in this transaction
        supplier_n = len(supplier_order)

        # Pad unused STS slots with empty values for uniform columns
        for i in range(supplier_n + 1, max_sts + 1):
            sts_fields[f"supplier{i}"]      = ""
            sts_fields[f"start_STS{i}"]     = ""
            sts_fields[f"starthour_STS{i}"] = ""
            sts_fields[f"end_STS{i}"]       = ""
            sts_fields[f"endhour_STS{i}"]   = ""
            sts_fields[f"duration_STS{i}"]  = ""

        # Total STS duration = sum of all aggregated supplier durations
        sts_fields["duration_STS"] = round(sts_total_duration)

        # ── Port extraction ────────────────────────────────────────────
        # PDF spec: port values must come from the reference table
        # {条帚门, 秀山东, 虾峙门, 衢山, 马峙}.
        # User requirement: if multiple ports appear (some matched,
        # some not), use the matched one.
        # Strategy: scan anchorage rows for the first matched port;
        # simultaneously collect unmatched location_details for
        # reporting when no match is found at all.
        port = ""
        unmatched_list = []
        anch_rows = grp[grp["is_anchorage"] == 1]
        for _, arow in anch_rows.iterrows():
            ld = arow.get("location_details")
            ld_str = str(ld).strip() if pd.notna(ld) else ""
            pm = arow.get("port_matched")
            pm_str = str(pm).strip() if pd.notna(pm) and str(pm).strip() not in ("", "nan") else ""
            if pm_str:
                port = pm_str
                break                              # First match wins (priority order pre-determined)
            if ld_str:
                unmatched_list.append(ld_str)

        # unmatched_port is only populated when NO reference port matched,
        # allowing manual review of unmatched transactions
        unmatched_port = "; ".join(unmatched_list) if not port and unmatched_list else ""

        # ── Draught ────────────────────────────────────────────────────
        # Mean draught across all rows in the transaction
        draught_vals = grp["draught_numeric"].dropna()
        draught = round(draught_vals.mean(), 2) if len(draught_vals) > 0 else 0.0

        # ── Assemble row ───────────────────────────────────────────────
        row = {
            "transaction_id": int(tid),
            **vessel_fields,
            "startdate":      startdate,
            "starthour":      starthour,
            "enddate":        enddate,
            "endhour":        endhour,
            "duration":       duration_total,
            "supplier_n":     supplier_n,
            **sts_fields,
            "port":           port,
            "unmatched_port": unmatched_port,
            "draught":        draught,
        }
        rows.append(row)

    result = pd.DataFrame(rows)
    print(f"  Built {len(result)} transaction rows")

    # ── 5. Deduplication ───────────────────────────────────────────────
    # PDF spec: remove duplicate transactions.  We compare all columns
    # except transaction_id (which is an artificial key).
    print("\nStep 5: Removing duplicates")
    before = len(result)
    dup_cols = [c for c in result.columns if c != "transaction_id"]
    result = result.drop_duplicates(subset=dup_cols)
    after = len(result)
    print(f"  Before: {before}, After: {after}, Removed: {before - after}")

    # ── 6. Save formatted final Excel ──────────────────────────────────
    # Column structure:
    #   [13 base columns] + [6 × max_sts STS columns] + [4 trailer columns]
    # = 13 + 6*max_sts + 4 columns total
    print(f"\nStep 6: Saving final file → {FINAL}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交易级别数据"

    # Build header list dynamically based on max_sts
    base_headers = [
        "transaction_id", "vessel_name", "vessel_code", "vessel_type", "vessel_status",
        "dwt", "gt",
        "startdate", "starthour", "enddate", "endhour", "duration",
        "supplier_n",
    ]
    sts_headers = []
    for i in range(1, max_sts + 1):
        sts_headers += [
            f"supplier{i}", f"start_STS{i}", f"starthour_STS{i}",
            f"end_STS{i}", f"endhour_STS{i}", f"duration_STS{i}",
        ]
    trailer_headers = ["duration_STS", "port", "unmatched_port", "draught"]
    all_headers = base_headers + sts_headers + trailer_headers

    # Apply header formatting
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for c, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Write data rows
    for i, (_, txn) in enumerate(result.iterrows()):
        r = i + 2
        base_vals = [
            txn.get("transaction_id"),
            txn.get("vessel_name"),
            txn.get("vessel_code"),
            txn.get("vessel_type"),
            txn.get("vessel_status"),
            txn.get("dwt"),
            txn.get("gt"),
            txn.get("startdate"),
            txn.get("starthour"),
            txn.get("enddate"),
            txn.get("endhour"),
            txn.get("duration"),
            txn.get("supplier_n"),
        ]
        sts_vals = []
        for j in range(1, max_sts + 1):
            sts_vals += [
                txn.get(f"supplier{j}"),
                txn.get(f"start_STS{j}"),
                txn.get(f"starthour_STS{j}"),
                txn.get(f"end_STS{j}"),
                txn.get(f"endhour_STS{j}"),
                txn.get(f"duration_STS{j}"),
            ]
        trailer_vals = [
            txn.get("duration_STS"),
            txn.get("port"),
            txn.get("unmatched_port"),
            txn.get("draught"),
        ]
        all_vals = base_vals + sts_vals + trailer_vals
        for c, val in enumerate(all_vals, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths: base columns + dynamic STS columns + trailer columns
    col_widths = {}
    # Base columns (1-13): transaction_id ~ supplier_n
    col_widths[1] = 14   # transaction_id
    col_widths[2] = 30   # vessel_name
    col_widths[3] = 12   # vessel_code
    col_widths[4] = 30   # vessel_type
    col_widths[5] = 24   # vessel_status
    col_widths[6] = 10   # dwt
    col_widths[7] = 10   # gt
    col_widths[8] = 16   # startdate
    col_widths[9] = 12   # starthour
    col_widths[10] = 16  # enddate
    col_widths[11] = 12  # endhour
    col_widths[12] = 12  # duration
    col_widths[13] = 12  # supplier_n
    # STS columns (14 through 13+6*max_sts): 6 columns per STS slot
    base_col = 14
    for i in range(1, max_sts + 1):
        col_widths[base_col + 0] = 18  # supplier{N}
        col_widths[base_col + 1] = 16  # start_STS{N}
        col_widths[base_col + 2] = 14  # starthour_STS{N}
        col_widths[base_col + 3] = 16  # end_STS{N}
        col_widths[base_col + 4] = 14  # endhour_STS{N}
        col_widths[base_col + 5] = 14  # duration_STS{N}
        base_col += 6
    # Trailer columns
    col_widths[base_col + 0] = 14  # duration_STS
    col_widths[base_col + 1] = 12  # port
    col_widths[base_col + 2] = 40  # unmatched_port
    col_widths[base_col + 3] = 12  # draught
    for col_idx, width in col_widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    wb.save(FINAL)
    print(f"  Saved {len(result)} rows × {len(all_headers)} columns")

    # ── 7. Validation: sample comparison ───────────────────────────────
    # Spot-check the first 5 transactions against the preprocessed source
    # to verify the aggregation logic is correct.
    print("\n" + "=" * 60)
    print("Step 7: Validation — sampling original vs processed")

    sample_tids = sorted(txns.groups.keys())[:5]
    for tid in sample_tids:
        grp = txns.get_group(tid)
        match_row = result[result["transaction_id"] == tid]
        if match_row.empty:
            print(f"  Txn {tid}: NOT FOUND in result!")
            continue
        processed = match_row.iloc[0]

        orig_sts = int(grp["is_sts"].sum())
        orig_anch = int(grp["is_anchorage"].sum())

        print(f"\n  Transaction {tid} ({processed['vessel_name']}):")
        print(f"    Original: {len(grp)} rows ({orig_sts} STS events, {orig_anch} Anchorage)")
        print(f"    Unique suppliers: {processed.get('supplier_n', 0)}")
        print(f"    Overall: {processed['startdate']} H{processed['starthour']} → "
              f"{processed['enddate']} H{processed['endhour']} ({processed['duration']}h)")

        processed_sts = int(processed.get("supplier_n", 0))
        for j in range(1, min(processed_sts + 1, 6)):
            print(f"      Supplier{j}: {processed[f'supplier{j}']} "
                  f"{processed[f'start_STS{j}']} H{processed[f'starthour_STS{j}']} → "
                  f"{processed[f'end_STS{j}']} H{processed[f'endhour_STS{j}']} "
                  f"({processed[f'duration_STS{j}']}h)")
        if processed_sts > 5:
            print(f"      ... and {processed_sts - 5} more suppliers")
        print(f"    STS total: {processed['duration_STS']}h")
        print(f"    Port: {processed['port']}")
        print(f"    Draught: {processed['draught']}")

    # ── 8. Global statistics ───────────────────────────────────────────
    print(f"\n  --- Global statistics ---")
    print(f"  Total transactions:       {len(result)}")

    # Multi-supplier: supplier_n >= 2
    multi_sts = result[result["supplier_n"] >= 2]
    print(f"  Multi-supplier transactions: {len(multi_sts)}")

    # Zero-STS: supplier_n == 0 (transaction has only anchorage rows)
    zero_sts = result[result["supplier_n"] == 0]
    print(f"  Zero-STS transactions:       {len(zero_sts)}")

    # Port match rate
    no_port = result[result["port"].isna() | result["port"].astype(str).str.strip().isin(["", "nan"])]
    port_matched = len(result) - len(no_port)
    print(f"  Port match rate:          {port_matched}/{len(result)} "
          f"({100 * port_matched / len(result):.1f}%)")

    port_dist = result["port"].fillna("No match").replace("", "No match").value_counts()
    print("  Port distribution:")
    for p, c in port_dist.items():
        print(f"    {p}: {c}")

    # Consistency: sum of all per-STS durations must equal duration_STS
    dur_cols = [f"duration_STS{i}" for i in range(1, max_sts + 1)]
    sts_dur_sum = pd.Series(0, index=result.index)
    for dc in dur_cols:
        sts_dur_sum += pd.to_numeric(result[dc], errors="coerce").fillna(0)
    mismatch = (result["duration_STS"].astype(float) != sts_dur_sum).sum()
    print(f"  duration_STS mismatch:    {mismatch}")

    neg_dur = (result["duration"] < 0).sum()
    print(f"  Negative overall duration: {neg_dur}")

    if len(no_port) > 0:
        no_port_ids = no_port["transaction_id"].tolist()
        print(f"\n  No-port transaction IDs (first 20): {no_port_ids[:20]}")
        if len(no_port) > 20:
            print(f"  ... {len(no_port)} total")

    print("\n" + "=" * 60)
    print("Finalization complete!")
    print(f"  Output: {FINAL}")


if __name__ == "__main__":
    main()
